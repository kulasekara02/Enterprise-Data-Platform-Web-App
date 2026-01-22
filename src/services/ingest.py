"""File ingestion service - handles file parsing and initial processing."""
import csv
import io
from typing import Generator, Dict, Any, List
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.api.config import settings


class IngestService:
    """Service for ingesting data files."""

    SUPPORTED_TYPES = {'csv', 'xlsx', 'xls', 'json'}

    def __init__(self, file_path: str, file_type: str):
        self.file_path = Path(file_path)
        self.file_type = file_type.lower()

        if self.file_type not in self.SUPPORTED_TYPES:
            raise ValueError(f"Unsupported file type: {file_type}")

    def get_row_count(self) -> int:
        """Get total row count without loading entire file into memory."""
        if self.file_type == 'csv':
            return self._count_csv_rows()
        elif self.file_type in ('xlsx', 'xls'):
            return self._count_excel_rows()
        elif self.file_type == 'json':
            return self._count_json_rows()
        return 0

    def _count_csv_rows(self) -> int:
        """Count rows in CSV file."""
        with open(self.file_path, 'r', encoding='utf-8') as f:
            return sum(1 for _ in f) - 1  # Subtract header

    def _count_excel_rows(self) -> int:
        """Count rows in Excel file."""
        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active
        count = ws.max_row - 1  # Subtract header
        wb.close()
        return count

    def _count_json_rows(self) -> int:
        """Count rows in JSON file."""
        df = pd.read_json(self.file_path)
        return len(df)

    def get_headers(self) -> List[str]:
        """Get column headers from file."""
        if self.file_type == 'csv':
            with open(self.file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                return next(reader)
        elif self.file_type in ('xlsx', 'xls'):
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            wb.close()
            return headers
        elif self.file_type == 'json':
            df = pd.read_json(self.file_path, nrows=1)
            return df.columns.tolist()
        return []

    def read_in_batches(self, batch_size: int = None) -> Generator[pd.DataFrame, None, None]:
        """Read file in batches for memory-efficient processing."""
        batch_size = batch_size or settings.BATCH_SIZE

        if self.file_type == 'csv':
            yield from self._read_csv_batches(batch_size)
        elif self.file_type in ('xlsx', 'xls'):
            yield from self._read_excel_batches(batch_size)
        elif self.file_type == 'json':
            yield from self._read_json_batches(batch_size)

    def _read_csv_batches(self, batch_size: int) -> Generator[pd.DataFrame, None, None]:
        """Read CSV in batches."""
        for chunk in pd.read_csv(self.file_path, chunksize=batch_size):
            yield chunk

    def _read_excel_batches(self, batch_size: int) -> Generator[pd.DataFrame, None, None]:
        """Read Excel in batches."""
        # Excel doesn't support native chunking, so we load and split
        df = pd.read_excel(self.file_path)
        for i in range(0, len(df), batch_size):
            yield df.iloc[i:i + batch_size]

    def _read_json_batches(self, batch_size: int) -> Generator[pd.DataFrame, None, None]:
        """Read JSON in batches."""
        df = pd.read_json(self.file_path)
        for i in range(0, len(df), batch_size):
            yield df.iloc[i:i + batch_size]

    def read_all(self) -> pd.DataFrame:
        """Read entire file into DataFrame."""
        if self.file_type == 'csv':
            return pd.read_csv(self.file_path)
        elif self.file_type in ('xlsx', 'xls'):
            return pd.read_excel(self.file_path)
        elif self.file_type == 'json':
            return pd.read_json(self.file_path)
        raise ValueError(f"Cannot read file type: {self.file_type}")

    def preview(self, rows: int = 10) -> List[Dict[str, Any]]:
        """Get preview of file data."""
        if self.file_type == 'csv':
            df = pd.read_csv(self.file_path, nrows=rows)
        elif self.file_type in ('xlsx', 'xls'):
            df = pd.read_excel(self.file_path, nrows=rows)
        elif self.file_type == 'json':
            df = pd.read_json(self.file_path)
            df = df.head(rows)
        else:
            return []

        return df.to_dict(orient='records')


def detect_delimiter(file_content: bytes) -> str:
    """Detect CSV delimiter from file content."""
    sample = file_content[:4096].decode('utf-8', errors='ignore')

    delimiters = [',', ';', '\t', '|']
    counts = {d: sample.count(d) for d in delimiters}

    return max(counts, key=counts.get)


def detect_encoding(file_path: str) -> str:
    """Detect file encoding."""
    import chardet

    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read(10000))

    return result['encoding'] or 'utf-8'
